#!/usr/bin/env python3
"""
Match Score — measures how well the candidate's profile fits a specific job.

Formula (weighted, 100 points max):

  1. Domain match (40 pts max)
     - Weighted by domain importance. Candidate's domains (AI/ML, Healthcare, Finance,
       Enterprise) add points. Domains candidate lacks (Cybersecurity, E-commerce,
       Infrastructure, Mobile) reduce the score by being in the denominator.
     - Score = matched_domain_weight / total_domain_weight_in_job × 40
     
  2. Skill keyword match (20 pts max)
     - 7 categories, each worth ~2.9 pts
     - Score = matched categories / total categories found in job × 20
  
  3. Seniority fit (15 pts max)
     - Senior/Staff/Principal/Director: 15 pts (candidate is senior, 15+ yrs)
     - Mid-level (PM II, Associate): 8 pts
     - Intern/Junior: 0 pts
  
  4. Years requirement (10 pts max)
     - <=8 years required: 10 pts (candidate exceeds)
     - 9-12 years: 8 pts
     - 13-15 years: 5 pts
     - 16+ years: 2 pts (stretch)
     - Not specified: 7 pts
  
  5. Location fit (10 pts max)
     - Boston/MA: 10 pts
     - Remote/US: 8 pts
     - Other US city: 3 pts
     - Non-US: 0 pts
  
  6. Entrepreneurial fit (5 pts max)
     - Job mentions startup, 0-to-1, MVP, incubation: 5 pts (candidate is co-founder)
     - Not mentioned: 0 pts

Total = sum of all components, capped at 100.
"""

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
JOURNAL = BASE_DIR / "journal.xlsx"
PROFILE_PATH = BASE_DIR / "resume" / "MASTER_PROFILE.md"
DESCS_PATH = BASE_DIR / "output" / "descriptions_2026-07-14.json"
CONFIG_PATH = BASE_DIR / "config.json"

# Column indexes from the shared journal definition (single source of truth)
from journal import JOBS_COL_INDEX

# ─── Load domain definitions from config.json ───

def _load_domains():
    with open(CONFIG_PATH) as f:
        config = json.load(f)
    domains = {}
    for name, info in config.get("scoring", {}).get("domains", {}).items():
        domains[name] = {
            "weight": info["weight"],
            "has_in_profile": info["has_in_profile"],
            "keywords": info["keywords"],
        }
    return domains

DOMAINS = _load_domains()

# ─── Skill categories (same as ATS but for match scoring) ───

SKILL_CATEGORIES = {
    "Product Management": [
        "product manager", "product management", "product strategy", "product roadmap",
        "product lifecycle", "product vision", "product owner", "backlog",
        "prd", "user stories", "prioritization", "go-to-market", "discovery", "delivery",
    ],
    "AI/ML": [
        "ai", "ml", "machine learning", "genai", "llm", "artificial intelligence",
        "model accuracy", "ai product", "agentic", "ai agent", "ai platform",
        "generative ai", "classification model",
    ],
    "Healthcare Domain": [
        "healthcare", "ehr", "hl7", "fhir", "clinical", "patient", "medical",
        "pharma", "biotech", "medicare", "pharmacy", "ambulatory",
    ],
    "Methodology": [
        "agile", "scrum", "kanban", "sprint", "cross-functional", "stakeholder",
        "mvp", "okr", "kpi", "metrics", "hypothesis",
    ],
    "Technical Fluency": [
        "api", "cloud", "aws", "platform", "integration", "python", "architecture",
        "saas", "b2b", "data fabric", "event-driven",
    ],
    "Data & Analytics": [
        "data-driven", "analytics", "experiment", "hypothesis", "success criteria",
        "dashboard", "reporting", "ontology", "data model", "sentiment analysis",
    ],
    "Leadership": [
        "lead", "director", "principal", "staff", "mentor", "cross-cultural",
        "distributed team", "stakeholder management", "co-founder",
    ],
    "Certifications & Qualifications": [
        "aipmm", "certified product manager", "safe", "popm", "pmp",
        "cfa", "mba", "m.sc", "master of science", "bachelor",
        "ai literacy", "product owner/product manager", "product owner",
    ],
}


def score_domain_match(desc_lower, profile_lower):
    """Score domain alignment (40 pts max)."""
    matched_weight = 0
    total_weight = 0
    matched_domains = []
    missing_domains = []

    import re
    for domain_name, info in DOMAINS.items():
        weight = info["weight"]
        has_in_profile = info["has_in_profile"]
        keywords = info["keywords"]

        job_has = any(re.search(r'\b' + re.escape(kw) + r'\b', desc_lower) for kw in keywords)

        if job_has:
            total_weight += weight
            if has_in_profile:
                matched_weight += weight
                matched_domains.append(domain_name)
            else:
                missing_domains.append(domain_name)

    if total_weight == 0:
        return 0, matched_domains, missing_domains

    score = int(matched_weight / total_weight * 40)
    return score, matched_domains, missing_domains


def score_skill_match(desc_lower, profile_lower):
    """Score skill keyword overlap (25 pts max)."""
    matched = 0
    total = 0

    for category, keywords in SKILL_CATEGORIES.items():
        job_has = any(kw in desc_lower for kw in keywords)
        profile_has = any(kw in profile_lower for kw in keywords)

        if job_has:
            total += 1
            if profile_has:
                matched += 1

    if total == 0:
        return 0

    return int(matched / total * 20)


def score_certification_gap(desc_lower, profile_lower):
    """Deduct points for certifications required in JD but missing from profile.
    Returns negative points (0 to -15)."""
    CERT_KEYWORDS = [
        "cfa", "pmp", "aipmm", "safe", "popm", "cpa", "cissp",
        "aws certified", "azure certified", "google cloud certified",
        "phr", "shrm", "six sigma", "itil",
    ]
    deductions = 0
    for cert in CERT_KEYWORDS:
        if cert in desc_lower and cert not in profile_lower:
            deductions -= 5
    return max(-15, deductions)


def score_seniority(title_lower):
    """Score seniority fit (15 pts max)."""
    SENIOR_TERMS = ["senior", "staff", "principal", "lead", "director", "head of",
                    "vp", "group product", "chief", "sr.", "sr "]
    MID_TERMS = ["product manager ii", "product manager 2", "associate product"]

    if any(term in title_lower for term in SENIOR_TERMS):
        return 15
    elif any(term in title_lower for term in MID_TERMS):
        return 8
    elif "intern" in title_lower:
        return 0
    else:
        # Generic "Product Manager" — neutral, likely mid-level
        return 8


def score_years_requirement(desc_lower):
    """Score years requirement fit (10 pts max)."""
    years_match = re.search(r'(\d+)\+?\s*years?', desc_lower)
    if not years_match:
        return 7  # not specified — neutral

    years = int(years_match.group(1))
    if years <= 8:
        return 10
    elif years <= 12:
        return 8
    elif years <= 15:
        return 5
    else:
        return 2  # 16+ years — stretch


def score_location(location_lower, preferred_location=None):
    if preferred_location is None:
        config_path = Path(__file__).parent.parent / "config.json"
        with open(config_path) as f:
            config = json.load(f)
        preferred_location = config.get("scoring", {}).get("preferred_location", "")
        if not preferred_location:
            raise ValueError("scoring.preferred_location not found in config.json")
    
    pref_lower = preferred_location.lower()
    pref_parts = pref_lower.split()
    
    # Check for preferred location match
    if any(part in location_lower for part in pref_parts) or ", ma" in location_lower:
        return 10
    elif "remote" in location_lower or "united states" in location_lower:
        return 8
    else:
        return 3  # default for US locations


def score_entrepreneurship(desc_lower):
    """Score entrepreneurial fit (5 pts max)."""
    ENTREPRENEUR_KEYWORDS = ["startup", "0 to 1", "zero to one", "mvp", "incubation",
                              "new product", "from scratch", "launch", "co-founder",
                              "founding", "early stage", "seed"]
    if any(kw in desc_lower for kw in ENTREPRENEUR_KEYWORDS):
        return 5
    return 0


def calculate_match_score(job, description, profile_text):
    """
    Calculate match score for a job against the candidate's profile.
    Returns: (score 0-100, details dict)
    """
    desc_lower = description.lower()
    profile_lower = profile_text.lower()
    title_lower = job.get("title", "").lower()
    location_lower = job.get("location", "").lower()

    # 1. Domain match (40 pts)
    domain_score, matched_domains, missing_domains = score_domain_match(desc_lower, profile_lower)

    # 2. Skill match (25 pts)
    skill_score = score_skill_match(desc_lower, profile_lower)

    # 3. Seniority (15 pts)
    seniority_score = score_seniority(title_lower)

    # 4. Years requirement (10 pts)
    years_score = score_years_requirement(desc_lower)

    # 5. Location (10 pts)
    location_score = score_location(location_lower)

    # 6. Entrepreneurship (5 pts)
    entre_score = score_entrepreneurship(desc_lower)

    # 7. Certification gap (0 to -15 pts)
    cert_penalty = score_certification_gap(desc_lower, profile_lower)

    total = domain_score + skill_score + seniority_score + years_score + location_score + entre_score + cert_penalty
    total = max(0, min(100, total))

    details = {
        "domain_score": domain_score,
        "skill_score": skill_score,
        "seniority_score": seniority_score,
        "years_score": years_score,
        "location_score": location_score,
        "entrepreneurship_score": entre_score,
        "certification_penalty": cert_penalty,
        "matched_domains": matched_domains,
        "missing_domains": missing_domains,
    }

    return total, details


def main():
    with open(PROFILE_PATH) as f:
        profile = f.read()

    with open(DESCS_PATH) as f:
        descriptions = json.load(f)

    wb = load_workbook(str(JOURNAL))
    ws = wb["Jobs"]

    MATCH_COL = JOBS_COL_INDEX["match_score"]
    URL_COL = JOBS_COL_INDEX["url"]
    TITLE_COL = JOBS_COL_INDEX["title"]
    COMPANY_COL = JOBS_COL_INDEX["company"]
    LOCATION_COL = JOBS_COL_INDEX["location"]

    results = []
    updated = 0

    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=URL_COL).value
        if not url:
            continue
        url = str(url).strip()

        desc = descriptions.get(url, "")
        if not desc or len(desc) < 50:
            continue

        job = {
            "title": ws.cell(row=row_idx, column=TITLE_COL).value or "",
            "company": ws.cell(row=row_idx, column=COMPANY_COL).value or "",
            "location": ws.cell(row=row_idx, column=LOCATION_COL).value or "",
        }

        score, details = calculate_match_score(job, desc, profile)
        ws.cell(row=row_idx, column=MATCH_COL, value=score)
        updated += 1
        results.append((score, job["company"], job["title"], job["location"], details))

    wb.save(str(JOURNAL))

    # Print results
    from collections import Counter
    buckets = Counter()
    for r in results:
        b = r[0] // 10 * 10
        buckets[b] += 1

    print(f"Updated {updated} match scores in journal\n")
    print("Match Score Distribution:")
    for b in sorted(buckets.keys()):
        print(f"  {b}-{b+9}: {buckets[b]}")

    results.sort(key=lambda x: x[0], reverse=True)
    print(f"\nTop 20:")
    print(f"{'Score':>5} {'Domain':>6} {'Skill':>5} {'Sen':>4} {'Yrs':>4} {'Loc':>4} {'Ent':>4}  {'Company':<20} {'Title':<30}")
    print(f"{'-'*5} {'-'*6} {'-'*5} {'-'*4} {'-'*4} {'-'*4} {'-'*4}  {'-'*20} {'-'*30}")
    for s in results[:20]:
        d = s[4]
        print(f"{s[0]:>5} {d['domain_score']:>6} {d['skill_score']:>5} {d['seniority_score']:>4} {d['years_score']:>4} {d['location_score']:>4} {d['entrepreneurship_score']:>4}  {s[1][:19]:<20} {s[2][:29]:<30}")

    print(f"\nBottom 10:")
    for s in results[-10:]:
        d = s[4]
        print(f"{s[0]:>5} {d['domain_score']:>6} {d['skill_score']:>5} {d['seniority_score']:>4} {d['years_score']:>4} {d['location_score']:>4} {d['entrepreneurship_score']:>4}  {s[1][:19]:<20} {s[2][:29]:<30}")

    avg = sum(r[0] for r in results) / len(results)
    print(f"\nAverage: {avg:.1f}")

    # Priority breakdown — read thresholds from config
    import json as _json
    try:
        with open("config.json") as _f:
            _cfg = _json.load(_f)
        _high_thresh = _cfg.get("scoring", {}).get("priority_thresholds", {}).get("high", 80)
        _med_thresh = _cfg.get("scoring", {}).get("priority_thresholds", {}).get("medium", 65)
    except:
        _high_thresh, _med_thresh = 80, 65

    high = sum(1 for r in results if r[0] >= _high_thresh)
    med = sum(1 for r in results if _med_thresh <= r[0] < _high_thresh)
    low = sum(1 for r in results if r[0] < _med_thresh)
    print(f"Priority: High={high} | Medium={med} | Low={low}")


if __name__ == "__main__":
    main()