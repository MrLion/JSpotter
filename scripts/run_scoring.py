#!/usr/bin/env python3
"""
Full pipeline: fetch descriptions for unscored jobs → calculate scores → update journal.

Run after adding new jobs to journal.xlsx via journal.py --add.
This script finds jobs with no match score, fetches their descriptions, 
and calculates all scoring fields.

Usage:
  python3 run_scoring.py
"""

import json
import re
import subprocess
import sys
import time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("openpyxl required")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
JOURNAL = BASE_DIR / "journal.xlsx"
PROFILE_PATH = BASE_DIR / "resume" / "MASTER_PROFILE.md"
DESCS_DIR = BASE_DIR / "output"
CONFIG_PATH = BASE_DIR / "config.json"

# Import scoring algorithms
sys.path.insert(0, str(BASE_DIR))
from ats_score import calculate_ats_score
from match_score import calculate_match_score
from interview_prob import calculate_interview_prob

PRIORITY_COLORS = {
    'High': PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid'),
    'Medium': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'Low': PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid'),
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "en-US,en;q=0.9",
}


def fetch_job_description(url, timeout=20):
    try:
        result = subprocess.run(
            ["curl", "-s", "-L", "--max-time", str(timeout),
             "-H", f"User-Agent: {HEADERS['User-Agent']}",
             "-H", f"Accept: {HEADERS['Accept']}",
             "-H", f"Accept-Language: {HEADERS['Accept-Language']}",
             url],
            capture_output=True, text=True, timeout=timeout + 5
        )
        html = result.stdout
        if not html or len(html) < 1000:
            return None

        match = re.search(
            r'<div class="description__text[^"]*"[^>]*>(.*?)</div>\s*</section>',
            html, re.DOTALL
        )
        if match:
            text = re.sub(r'<[^>]+>', ' ', match.group(1))
            text = re.sub(r'\s+', ' ', text).strip()
            text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
            if len(text) > 50:
                return text

        # Fallback: JSON-LD
        match2 = re.search(r'"description"\s*:\s*"((?:[^"\\]|\\.)*)"', html)
        if match2:
            text = match2.group(1).replace('\\n', ' ').replace('\\"', '"')
            text = re.sub(r'<[^>]+>', ' ', text)
            text = re.sub(r'\\u003c[^>]*\\u003e', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            if len(text) > 50:
                return text

        return None
    except Exception:
        return None


def main():
    # Load profile
    with open(PROFILE_PATH) as f:
        profile = f.read()

    # Load existing descriptions cache
    desc_path = DESCS_DIR / "descriptions_cache.json"
    descriptions = {}
    if desc_path.exists():
        with open(desc_path) as f:
            descriptions = json.load(f)
    print(f"Loaded {len(descriptions)} cached descriptions")

    # Load journal and find unscored jobs
    wb = load_workbook(str(JOURNAL))
    ws = wb["Jobs"]

    COL = {
        "date_found": 1, "company": 2, "title": 3, "location": 4,
        "source": 5, "url": 6, "match_score": 7, "ats_score": 8,
        "missing_skills": 9, "salary_estimate": 10, "interview_prob": 11,
        "fit_notes": 12, "priority": 13, "job_id": 14, "status": 15,
        "app_url": 16, "date_applied": 17,
    }

    unscored = []
    for row_idx in range(2, ws.max_row + 1):
        match_score = ws.cell(row=row_idx, column=COL["match_score"]).value
        url = ws.cell(row=row_idx, column=COL["url"]).value
        if not url:
            continue
        url = str(url).strip()
        if match_score is None:
            unscored.append({
                "row": row_idx,
                "url": url,
                "title": ws.cell(row=row_idx, column=COL["title"]).value or "",
                "company": ws.cell(row=row_idx, column=COL["company"]).value or "",
                "location": ws.cell(row=row_idx, column=COL["location"]).value or "",
            })

    print(f"Found {len(unscored)} unscored jobs")

    if not unscored:
        print("All jobs already scored. Nothing to do.")
        wb.close()
        return

    # Fetch descriptions for unscored jobs that aren't cached
    to_fetch = [j for j in unscored if j["url"] not in descriptions]
    print(f"Need to fetch {len(to_fetch)} descriptions...")

    if to_fetch:
        with ThreadPoolExecutor(max_workers=2) as executor:
            future_to_job = {
                executor.submit(fetch_job_description, j["url"]): j
                for j in to_fetch
            }
            fetched = 0
            for future in as_completed(future_to_job):
                job = future_to_job[future]
                try:
                    desc = future.result()
                    if desc and job["url"] not in descriptions:
                        descriptions[job["url"]] = desc
                        fetched += 1
                except Exception:
                    pass
                if (fetched + 1) % 5 == 0:
                    print(f"  Fetched {fetched}/{len(to_fetch)}...")

        # Retry failed ones
        still_missing = [j for j in to_fetch if j["url"] not in descriptions]
        if still_missing:
            print(f"  Retrying {len(still_missing)} failed fetches...")
            for j in still_missing:
                desc = fetch_job_description(j["url"], timeout=30)
                if desc and j["url"] not in descriptions:
                    descriptions[j["url"]] = desc
                time.sleep(2)  # slow pacing to avoid LinkedIn rate limits

        # Save updated cache
        with open(desc_path, "w") as f:
            json.dump(descriptions, f, indent=2)
        print(f"  Total descriptions cached: {len(descriptions)}")

    # Score each unscored job
    print(f"\nScoring {len(unscored)} jobs...")
    scored = 0

    # Priority thresholds — read from config once
    try:
        with open(CONFIG_PATH) as _f:
            _cfg = json.load(_f)
        _high_thresh = _cfg.get("scoring", {}).get("priority_thresholds", {}).get("high", 80)
        _med_thresh = _cfg.get("scoring", {}).get("priority_thresholds", {}).get("medium", 65)
    except Exception:
        _high_thresh, _med_thresh = 80, 65

    for job in unscored:
        desc = descriptions.get(job["url"], "")
        row_idx = job["row"]

        if not desc or len(desc) < 50:
            # Can't score without description
            ws.cell(row=row_idx, column=COL["match_score"], value=0)
            ws.cell(row=row_idx, column=COL["ats_score"], value=0)
            ws.cell(row=row_idx, column=COL["interview_prob"], value="5%")
            ws.cell(row=row_idx, column=COL["priority"], value="Low")
            ws.cell(row=row_idx, column=COL["priority"]).fill = PRIORITY_COLORS["Low"]
            ws.cell(row=row_idx, column=COL["fit_notes"], value="No description available")
            continue

        # Calculate all scores
        ats_score, ats_details = calculate_ats_score(desc, profile)
        match_score, match_details = calculate_match_score(job, desc, profile)
        prob = calculate_interview_prob(match_score, ats_score, job["title"], desc, job["location"])

        if match_score >= _high_thresh:
            priority = "High"
        elif match_score >= _med_thresh:
            priority = "Medium"
        else:
            priority = "Low"

        # Missing skills from match details
        missing = match_details.get("missing_domains", [])
        missing_str = ", ".join(missing) if missing else ""

        # Build fit notes
        notes = []
        if match_details.get("matched_domains"):
            notes.append(f"Strong: {', '.join(match_details['matched_domains'][:3])}")
        if missing:
            notes.append(f"Gap: {', '.join(missing[:2])}")
        if not notes:
            notes.append("General PM fit")
        fit_notes = " | ".join(notes)

        # Write to journal
        ws.cell(row=row_idx, column=COL["match_score"], value=match_score)
        ws.cell(row=row_idx, column=COL["ats_score"], value=ats_score)
        ws.cell(row=row_idx, column=COL["missing_skills"], value=missing_str)
        ws.cell(row=row_idx, column=COL["interview_prob"], value=f"{prob}%")
        ws.cell(row=row_idx, column=COL["fit_notes"], value=fit_notes)
        ws.cell(row=row_idx, column=COL["priority"], value=priority)
        ws.cell(row=row_idx, column=COL["priority"]).fill = PRIORITY_COLORS[priority]
        
        # Apply text wrapping for long-text columns
        from openpyxl.styles import Alignment
        wrap = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row_idx, column=COL["missing_skills"]).alignment = wrap
        ws.cell(row=row_idx, column=COL["fit_notes"]).alignment = wrap

        # Default status
        if not ws.cell(row=row_idx, column=COL["status"]).value:
            ws.cell(row=row_idx, column=COL["status"], value="Not Applied")

        scored += 1
        print(f"  {job['company'][:20]:<20} match={match_score:>3} ats={ats_score:>3} prob={prob:>2}%  {priority:<7} {job['title'][:30]}")

    # Ensure auto-filters on all sheets
    from openpyxl.utils import get_column_letter
    for sheet_name in ['Jobs', 'Applications', 'Resume Versions']:
        ws_filter = wb[sheet_name]
        max_col = ws_filter.max_column
        ws_filter.auto_filter.ref = f'A1:{get_column_letter(max_col)}1'
    
    # Color-code company cells by status
    from openpyxl.styles import PatternFill, Font
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    blue_fill = PatternFill(start_color='B3D9FF', end_color='B3D9FF', fill_type='solid')
    
    for row in ws.iter_rows(min_row=2, max_col=20):
        status_cell = ws.cell(row=row[0].row, column=15)
        status = str(status_cell.value).lower() if status_cell.value else ''
        company_cell = row[1]
        
        if 'interview' in status:
            company_cell.fill = blue_fill
            company_cell.font = Font(bold=True)
        elif 'applied' in status and 'not' not in status:
            company_cell.fill = green_fill
            company_cell.font = Font(bold=True)
        elif 'reject' in status:
            company_cell.fill = red_fill
        elif 'closed' in status:
            company_cell.fill = grey_fill
        elif 'not applied' in status or status == '':
            company_cell.fill = yellow_fill
    
    wb.save(str(JOURNAL))
    print(f"Scored {scored} jobs. Journal updated.")

    # Summary
    from collections import Counter
    priorities = Counter()
    for job in unscored:
        desc = descriptions.get(job["url"], "")
        if not desc or len(desc) < 50:
            priorities["Low"] += 1
            continue
        ms, _ = calculate_match_score(job, desc, profile)
        if ms >= _high_thresh:
            priorities["High"] += 1
        elif ms >= _med_thresh:
            priorities["Medium"] += 1
        else:
            priorities["Low"] += 1

    print(f"\nNew jobs priority: High={priorities['High']} Medium={priorities['Medium']} Low={priorities['Low']}")


if __name__ == "__main__":
    main()