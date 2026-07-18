#!/usr/bin/env python3
"""
Interview Probability — estimates likelihood of getting an interview.

Based on factors a recruiter would weigh:

1. Match score (40% weight) — how well profile fits the job
2. ATS score (20% weight) — keyword alignment for resume screening
3. Missing skills penalty (15% weight) — each missing domain reduces odds
4. Years requirement (10% weight) — stretch requirements reduce odds
5. Seniority alignment (10% weight) — over/under qualification
6. Location (5% weight) — local candidates preferred

Formula: base_probability × adjustment_factors, clamped 5-95%
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
JOURNAL = BASE_DIR / "journal.xlsx"
PROFILE_PATH = BASE_DIR / "resume" / "MASTER_PROFILE.md"
DESCS_PATH = BASE_DIR / "output" / "descriptions_2026-07-14.json"

# Domains the candidate lacks (same as match_score.py)
MISSING_DOMAINS = ["Cybersecurity", "E-commerce/Consumer", "Infrastructure/DevOps", "Mobile/Hardware"]

DOMAIN_KEYWORDS = {
    "Cybersecurity": ["security", "cyber", "threat", "vulnerability", "encryption", "soc", "siem",
                      "compliance", "fraud", "identity"],
    "E-commerce/Consumer": ["ecommerce", "e-commerce", "marketplace", "retail", "b2c", "shopping",
                             "consumer", "loyalty", "brand"],
    "Infrastructure/DevOps": ["infrastructure", "devops", "kubernetes", "docker", "sre",
                              "observability", "monitoring", "reliability", "scalability"],
    "Mobile/Hardware": ["mobile", "ios", "android", "hardware", "embedded", "firmware",
                        "device", "iot", "sensors"],
}


def calculate_interview_prob(match_score, ats_score, title, description, location):
    """
    Calculate interview probability (5-95%).
    
    Returns: integer percentage
    """
    desc_lower = description.lower()
    title_lower = title.lower()
    loc_lower = location.lower()

    # ── 1. Base from match score (40% weight) ──
    # Match score 80+ → strong base, 60-79 → moderate, <60 → weak
    if match_score >= 85:
        base = 70
    elif match_score >= 75:
        base = 60
    elif match_score >= 65:
        base = 48
    elif match_score >= 55:
        base = 35
    else:
        base = 20

    # ── 2. ATS score adjustment (20% weight) ──
    # High ATS = resume will pass automated screening
    # Range: -10 to +15
    if ats_score >= 90:
        ats_adj = 15
    elif ats_score >= 80:
        ats_adj = 10
    elif ats_score >= 70:
        ats_adj = 5
    elif ats_score >= 60:
        ats_adj = 0
    else:
        ats_adj = -10

    # ── 3. Missing skills penalty (15% weight) ──
    # Count how many domains the candidate lacks that the job requires
    missing_count = 0
    for domain, keywords in DOMAIN_KEYWORDS.items():
        if any(kw in desc_lower for kw in keywords):
            missing_count += 1

    # Each missing domain reduces probability
    missing_adj = missing_count * -6  # -6 per missing domain (max -24)

    # ── 4. Years requirement (10% weight) ──
    years_match = re.search(r'(\d+)\+?\s*years?', desc_lower)
    if years_match:
        years = int(years_match.group(1))
        if years <= 8:
            years_adj = 8   # candidate exceeds requirement
        elif years <= 12:
            years_adj = 4
        elif years <= 15:
            years_adj = -2
        else:
            years_adj = -15  # 16+ years — significant stretch
    else:
        years_adj = 3  # not specified — slight positive

    # ── 5. Seniority alignment (10% weight) ──
    SENIOR_TERMS = ["senior", "staff", "principal", "lead", "director", "head of", "vp", "group product", "chief"]
    MID_TERMS = ["product manager ii", "product manager 2", "associate product"]
    
    is_senior = any(term in title_lower for term in SENIOR_TERMS)
    is_mid = any(term in title_lower for term in MID_TERMS)
    is_junior = "intern" in title_lower

    if is_senior:
        sen_adj = 8   # candidate is senior, fits senior roles
    elif is_mid:
        sen_adj = 3   # candidate overqualified but still viable
    elif is_junior:
        sen_adj = -12  # significant overqualification — recruiter concern
    else:
        sen_adj = 2   # generic PM role — neutral-positive

    # ── 6. Location (5% weight) ──
    if "boston" in loc_lower or ", ma" in loc_lower or "mass" in loc_lower:
        loc_adj = 5   # local candidate — recruiter prefers
    elif "remote" in loc_lower or "united states" in loc_lower:
        loc_adj = 3   # remote — no location barrier
    else:
        loc_adj = -3  # relocation needed — reduces odds

    # ── Calculate final ──
    prob = base + ats_adj + missing_adj + years_adj + sen_adj + loc_adj
    prob = max(5, min(95, int(prob)))

    return prob


def main():
    with open(PROFILE_PATH) as f:
        profile = f.read()
    with open(DESCS_PATH) as f:
        descriptions = json.load(f)

    wb = load_workbook(str(JOURNAL))
    ws = wb["Jobs"]

    MATCH_COL = 7
    ATS_COL = 8
    PROB_COL = 11
    COMPANY_COL = 2
    TITLE_COL = 3
    LOCATION_COL = 4
    URL_COL = 6

    results = []
    updated = 0

    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=URL_COL).value
        if not url:
            continue
        url = str(url).strip()

        match_score = ws.cell(row=row_idx, column=MATCH_COL).value or 0
        ats_score = ws.cell(row=row_idx, column=ATS_COL).value or 0
        title = ws.cell(row=row_idx, column=TITLE_COL).value or ""
        location = ws.cell(row=row_idx, column=LOCATION_COL).value or ""
        company = ws.cell(row=row_idx, column=COMPANY_COL).value or ""

        desc = descriptions.get(url, "")
        if not desc or len(desc) < 50:
            continue

        prob = calculate_interview_prob(match_score, ats_score, title, desc, location)
        ws.cell(row=row_idx, column=PROB_COL, value=f"{prob}%")
        updated += 1
        results.append((prob, company, title, match_score, ats_score))

    wb.save(str(JOURNAL))

    # Print results
    from collections import Counter
    buckets = Counter()
    for r in results:
        b = r[0] // 10 * 10
        buckets[b] += 1

    print(f"Updated {updated} interview probabilities in journal\n")
    print("Distribution:")
    for b in sorted(buckets.keys()):
        print(f"  {b}-{b+9}%: {buckets[b]}")

    results.sort(key=lambda x: x[0], reverse=True)
    print(f"\nTop 10:")
    print(f"{'Prob':>5} {'Match':>5} {'ATS':>4} {'Company':<22} {'Title':<30}")
    print(f"{'-'*5} {'-'*5} {'-'*4} {'-'*22} {'-'*30}")
    for r in results[:10]:
        print(f"{r[0]:>4}% {r[3]:>5} {r[4]:>4} {r[1][:21]:<22} {r[2][:29]:<30}")

    print(f"\nBottom 10:")
    for r in results[-10:]:
        print(f"{r[0]:>4}% {r[3]:>5} {r[4]:>4} {r[1][:21]:<22} {r[2][:29]:<30}")

    avg = sum(r[0] for r in results) / len(results)
    print(f"\nAverage: {avg:.1f}%")


if __name__ == "__main__":
    main()