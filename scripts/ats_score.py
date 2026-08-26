#!/usr/bin/env python3
"""
ATS Scoring — measures keyword overlap between job description and candidate's master profile.

How it works:
1. Scan each job description for a defined set of ATS keywords (grouped by category)
2. For each keyword found in the job, check if it also appears in the candidate's profile
3. ATS score = (keywords in job AND in profile) / (keywords in job) × 100

This measures: "Of the keywords this job posting asks for, what percentage does the
candidate's resume contain?" — which is what an ATS actually checks.

Updates journal.xlsx with recalculated ATS scores.
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
JOURNAL = BASE_DIR / "journal.xlsx"
PROFILE_PATH = BASE_DIR / "resume" / "MASTER_PROFILE.md"
DESCS_PATH = BASE_DIR / "output" / "descriptions_2026-07-14.json"

# ATS keyword categories — these are the terms ATS systems scan for
# Each category is weighted equally (by keyword count within it)
ATS_CATEGORIES = {
    "Product Management": [
        "product manager", "product management", "product strategy", "product roadmap",
        "product lifecycle", "product vision", "product owner", "backlog management",
        "product backlog", "prd", "user stories", "prioritization", "go-to-market",
        "product discovery", "product delivery", "minimum viable product",
    ],
    "AI/ML": [
        "ai", "ml", "machine learning", "genai", "llm", "artificial intelligence",
        "ai/ml", "model accuracy", "ai product", "ai platform", "agentic", "ai agent",
        "ai features", "ai solutions", "ai enablement", "intelligent automation",
        "generative ai", "model training", "annotation", "classification model",
    ],
    "Healthcare": [
        "healthcare", "health", "ehr", "hl7", "fhir", "clinical", "patient",
        "medical", "pharma", "biotech", "payer", "provider", "medicare",
        "insurance", "ambulatory", "medical device", "wearable", "pharmacy",
    ],
    "Methodology": [
        "agile", "scrum", "kanban", "scrumban", "sprint", "cross-functional",
        "stakeholder", "discovery", "delivery", "mvp", "okr", "kpi", "metrics",
        "hypothesis", "user acceptance testing", "uat", "iteration",
    ],
    "Technical": [
        "api", "cloud", "aws", "platform", "integration", "data fabric",
        "event-driven", "python", "architecture", "infrastructure", "saas",
        "b2b", "b2c", "digital transformation", "low-code", "interoperability",
    ],
    "Data & Analytics": [
        "data-driven", "analytics", "experiment", "hypothesis",
        "a/b test", "success criteria", "dashboard", "reporting",
        "data model", "ontology", "data pipeline", "sentiment analysis",
    ],
    "Leadership": [
        "lead", "director", "principal", "staff", "mentor", "cross-cultural",
        "distributed team", "team alignment", "stakeholder management",
        "executive", "head of product", "co-founder",
    ],
}


def calculate_ats_score(job_description, profile_text):
    """
    Calculate ATS compatibility score.
    
    Returns: (score 0-100, details dict)
    """
    desc_lower = job_description.lower()
    profile_lower = profile_text.lower()

    total_keywords_found = 0
    matched_keywords = 0
    details = {}

    for category, keywords in ATS_CATEGORIES.items():
        cat_in_job = [kw for kw in keywords if kw in desc_lower]
        cat_in_both = [kw for kw in cat_in_job if kw in profile_lower]
        
        total_keywords_found += len(cat_in_job)
        matched_keywords += len(cat_in_both)
        
        details[category] = {
            "in_job": len(cat_in_job),
            "in_both": len(cat_in_both),
            "missing_from_profile": [kw for kw in cat_in_job if kw not in profile_lower],
        }

    if total_keywords_found == 0:
        return 0, details

    score = int(matched_keywords / total_keywords_found * 100)
    return score, details


def main():
    # Load profile
    with open(PROFILE_PATH) as f:
        profile = f.read()

    # Load descriptions
    with open(DESCS_PATH) as f:
        descriptions = json.load(f)

    # Load journal
    wb = load_workbook(str(JOURNAL))
    ws = wb["Jobs"]

    # Find URL column (column 6) and ATS score column (column 8)
    URL_COL = 6
    ATS_COL = 8

    updated = 0
    scores = []

    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=URL_COL).value
        if not url:
            continue
        url = str(url).strip()

        desc = descriptions.get(url, "")
        if not desc or len(desc) < 50:
            continue

        score, details = calculate_ats_score(desc, profile)
        ws.cell(row=row_idx, column=ATS_COL, value=score)
        updated += 1

        company = ws.cell(row=row_idx, column=2).value or ""
        title = ws.cell(row=row_idx, column=3).value or ""
        scores.append((score, company, title, details))

    wb.save(str(JOURNAL))

    # Print summary
    print(f"Updated {updated} ATS scores in journal\n")

    # Distribution
    from collections import Counter
    buckets = Counter()
    for s in scores:
        b = s[0] // 10 * 10
        buckets[b] += 1
    print("ATS Score Distribution:")
    for b in sorted(buckets.keys()):
        print(f"  {b}-{b+9}: {buckets[b]}")

    # Top 10
    scores.sort(key=lambda x: x[0], reverse=True)
    print(f"\nTop 10 ATS Scores:")
    print(f"{'ATS':>4} {'Company':<22} {'Title':<35}")
    print(f"{'-'*4} {'-'*22} {'-'*35}")
    for s in scores[:10]:
        print(f"{s[0]:>4} {s[1][:21]:<22} {s[2][:34]:<35}")

    # Bottom 5 with missing keywords
    print(f"\nBottom 5 ATS Scores (with gaps):")
    for s in scores[-5:]:
        missing = []
        for cat, d in s[3].items():
            missing.extend(d["missing_from_profile"][:3])
        print(f"  {s[0]:>3} {s[1][:20]}: missing {missing[:5]}")

    avg = sum(s[0] for s in scores) / len(scores)
    print(f"\nAverage ATS score: {avg:.1f}")


if __name__ == "__main__":
    main()