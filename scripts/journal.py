#!/usr/bin/env python3
"""
Job Journal — Excel-based application & resume tracking spreadsheet.

Creates and maintains an .xlsx journal with sheets for:
  1. Jobs — all discovered jobs with scoring fields
  2. Applications — tracking pipeline (applied, interview, offer, rejected)
  3. Resume Versions — tailored resume tracking per job

Usage:
  python3 journal.py --init                           # create empty journal
  python3 journal.py --add output/linkedin_extract.json  # add jobs from JSON
  python3 journal.py --remove --url 4449017604        # remove by job URL substring
  python3 journal.py --remove --company "Apple" --title "Product Manager"  # remove by company/title
  python3 journal.py --remove-all-medlow              # remove all Medium/Low Not-Applied jobs
  python3 journal.py --status                          # show summary stats
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
JOURNAL_PATH = BASE_DIR / "journal.xlsx"

# ─── Column definitions ───

JOBS_COLUMNS = [
    ("date_found", "Date Found", 12),
    ("company", "Company", 22),
    ("title", "Job Title", 35),
    ("location", "Location", 22),
    ("search_mode", "Source", 10),
    ("url", "Job URL", 45),
    ("match_score", "Match Score", 12),
    ("ats_score", "ATS Score", 12),
    ("missing_skills", "Missing Skills", 30),
    ("salary_estimate", "Salary Estimate", 16),
    ("fit_notes", "Fit Notes", 40),
    ("priority", "Priority", 10),

    ("job_id", "Job ID", 14),
    ("status", "Status", 16),
    ("app_url", "App URL", 45),
    ("date_applied", "Date Applied", 12),
    ("last_updated", "Last Updated", 12),
]

# 1-based column index per Jobs-sheet key — the single source of truth for
# every script that reads/writes the journal. Never hardcode a column number.
JOBS_COL_INDEX: dict[str, int] = {key: idx for idx, (key, _, _) in enumerate(JOBS_COLUMNS, 1)}

APPLICATIONS_COLUMNS = [
    ("date_found", "Date Found", 12),
    ("company", "Company", 22),
    ("title", "Job Title", 35),
    ("date_applied", "Date Applied", 12),
    ("resume_version", "Resume Version", 18),
    ("cover_letter", "Cover Letter", 12),
    ("status", "Status", 16),
    ("interview_date", "Interview Date", 14),
    ("interview_round", "Round", 8),
    ("offer_amount", "Offer Amount", 14),
    ("offer_date", "Offer Date", 12),
    ("rejected_date", "Rejected Date", 12),
    ("rejection_reason", "Rejection Reason", 30),
    ("notes", "Notes", 40),
    ("url", "Job URL", 45),
]

RESUME_COLUMNS = [
    ("date", "Date", 12),
    ("company", "Company", 22),
    ("title", "Job Title", 35),
    ("resume_file", "Resume File", 30),
    ("cover_letter_file", "Cover Letter File", 30),
    ("keywords_added", "Keywords Added", 30),
    ("keywords_removed", "Keywords Removed", 30),
    ("summary_used", "Summary Used", 50),
    ("highlights_changed", "Highlights Changed", 40),
    ("notes", "Notes", 40),
]

# Status values for Applications sheet
STATUS_VALUES = ["Not Applied", "Applied", "Phone Screen", "Interview", "Final Round", "Offer", "Rejected", "Withdrawn"]

# Status color coding for company cell
STATUS_COLORS = {
    "Not Applied": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # yellow
    "Applied": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),       # green
    "Phone Screen": PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid"),  # light blue
    "Interview": PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid"),     # blue
    "Final Round": PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid"),   # blue
    "Offer": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),         # green
    "Rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),      # red
    "Withdrawn": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),     # grey
}

# ─── Styling ───

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

PRIORITY_COLORS = {
    "High": PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid"),
    "Medium": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Low": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
}

SCORE_COLORS = {
    "high": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),   # green
    "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # yellow
    "low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),     # red
}


def init_journal(path=JOURNAL_PATH):
    """Create a fresh journal with all three sheets."""
    wb = Workbook()

    # ── Sheet 1: Jobs ──
    ws_jobs = wb.active
    ws_jobs.title = "Jobs"
    _write_headers(ws_jobs, JOBS_COLUMNS)
    _format_sheet(ws_jobs, JOBS_COLUMNS)

    # ── Sheet 2: Applications ──
    ws_apps = wb.create_sheet("Applications")
    _write_headers(ws_apps, APPLICATIONS_COLUMNS)
    _format_sheet(ws_apps, APPLICATIONS_COLUMNS)

    # ── Sheet 3: Resume Versions ──
    ws_resume = wb.create_sheet("Resume Versions")
    _write_headers(ws_resume, RESUME_COLUMNS)
    _format_sheet(ws_resume, RESUME_COLUMNS)

    # ── Sheet 4: Reference ──
    ws_ref = wb.create_sheet("Reference")
    ws_ref["A1"] = "Job Search Journal — Reference"
    ws_ref["A1"].font = Font(bold=True, size=14)
    ref_data = [
        ["", ""],
        ["Status Values (Applications sheet)", ""],
        ["Not Applied", "Job found but not yet applied"],
        ["Applied", "Application submitted"],
        ["Phone Screen", "Initial phone/screening call"],
        ["Interview", "Technical or on-site interview"],
        ["Final Round", "Final stage interview"],
        ["Offer", "Offer received"],
        ["Rejected", "Application rejected"],
        ["Withdrawn", "Withdrew application"],
        ["", ""],
        ["Priority Values (Jobs sheet)", ""],
        ["High", "Top match, apply within 24h"],
        ["Medium", "Good match, apply this week"],
        ["Low", "Possible match, monitor"],
        ["", ""],
        ["Search Mode (Jobs sheet)", ""],
        ["boston", "Found via Greater Boston LinkedIn search"],
        ["remote", "Found via Remote USA LinkedIn search"],
        ["greenhouse", "Found via Greenhouse ATS API"],
        ["lever", "Found via Lever ATS API"],
        ["ashby", "Found via Ashby ATS API"],
    ]
    for row in ref_data:
        ws_ref.append(row)
    ws_ref.column_dimensions["A"].width = 30
    ws_ref.column_dimensions["B"].width = 50

    wb.save(str(path))
    print(f"Journal created: {path}")
    return path


def _write_headers(ws, columns):
    """Write header row from column definitions."""
    for idx, (key, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _format_sheet(ws, columns):
    """Apply column widths, freeze panes, auto-filter, and text wrapping."""
    for idx, (key, label, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    # Enable text wrapping for all data cells
    from openpyxl.styles import Alignment
    wrap = Alignment(wrap_text=True, vertical='top')
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = wrap


def _load_adzuna_salaries():
    """Map adzuna job URLs -> '$135,624' for non-predicted salaries only.

    Annual-range sanity bounds come from config.json → hard_constraints →
    salary_sanity (defaults 50k–2M). Returns {} silently when the extract
    file is absent or unreadable.
    """
    path = BASE_DIR / "output" / "adzuna_extract.json"
    try:
        with open(path) as f:
            records = json.load(f)
        # Salary sanity bounds from config (fall back to defaults)
        try:
            with open(BASE_DIR / "config.json") as f:
                _hc = json.load(f).get("hard_constraints", {})
            _sanity = _hc.get("salary_sanity", {})
            _min = int(_sanity.get("min_annual", 50000))
            _max = int(_sanity.get("max_annual", 2_000_000))
        except Exception:
            _min, _max = 50000, 2_000_000
        out = {}
        for r in records:
            try:
                if str(r.get("salary_is_predicted")) == "0" and r.get("salary_min"):
                    lo, hi = int(r["salary_min"]), int(r.get("salary_max") or r["salary_min"])
                    # Sanity: annual USD range; drops hourly rates & typos
                    if _min <= lo <= _max and _min <= hi <= _max:
                        out[r["url"]] = f"${lo:,}–${hi:,}" if hi > lo else f"${lo:,}"
            except (KeyError, TypeError, ValueError):
                continue
        return out
    except Exception:
        return {}


def add_jobs(jobs_data, path=JOURNAL_PATH):
    """Add jobs to the Jobs sheet. Skips duplicates by App URL, then URL, then company+title+location."""
    if not path.exists():
        init_journal(path)

    wb = load_workbook(str(path))
    ws = wb["Jobs"]

    # Get existing entries for dedup
    url_col = _find_col_by_key(JOBS_COLUMNS, "url")
    job_id_col = _find_col_by_key(JOBS_COLUMNS, "job_id")
    app_url_col = _find_col_by_key(JOBS_COLUMNS, "app_url")
    company_col = _find_col_by_key(JOBS_COLUMNS, "company")
    title_col = _find_col_by_key(JOBS_COLUMNS, "title")
    location_col = _find_col_by_key(JOBS_COLUMNS, "location")

    existing_urls = set()
    existing_job_ids = set()
    existing_app_urls = set()
    existing_company_title_loc = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        url_val = row[url_col - 1].value
        if url_val:
            existing_urls.add(str(url_val).strip())
        job_id_val = row[job_id_col - 1].value
        if job_id_val:
            existing_job_ids.add(str(job_id_val).strip())
        app_url_val = row[app_url_col - 1].value if app_url_col <= len(row) else None
        if app_url_val:
            existing_app_urls.add(str(app_url_val).strip())
        company_val = row[company_col - 1].value
        title_val = row[title_col - 1].value
        location_val = row[location_col - 1].value
        if company_val and title_val:
            key = f"{str(company_val).strip().lower()}|{str(title_val).strip().lower()}|{str(location_val or '').strip().lower()}"
            existing_company_title_loc.add(key)

    # Find next empty row
    next_row = ws.max_row + 1
    if ws.max_row == 1:  # only headers
        next_row = 2

    added = 0
    skipped = 0
    date_str = datetime.now().strftime("%Y-%m-%d")

    # Real (non-predicted) Adzuna salaries for Salary Estimate autofill
    salary_by_url = _load_adzuna_salaries()

    for job in jobs_data:
        url = job.get("url", "").strip()
        app_url = job.get("app_url", "").strip()
        company = job.get("company", "").strip()
        title = job.get("title", "").strip()
        location = job.get("location", "").strip()

        # Dedup by App URL (highest priority — cross-source)
        if app_url and app_url in existing_app_urls:
            skipped += 1
            continue

        # Dedup by source URL
        if url in existing_urls:
            skipped += 1
            continue

        # Extract job ID from URL — only if it contains digits (real IDs have numbers)
        job_id = ""
        if url:
            last_segment = url.rstrip("/").split("/")[-1]
            if last_segment and any(c.isdigit() for c in last_segment):
                job_id = last_segment

        # Also check by job ID
        if job_id and job_id in existing_job_ids:
            skipped += 1
            continue

        # Fallback: dedup by company + title + location (catches cross-source dupes without App URL)
        if company and title:
            key = f"{company.lower()}|{title.lower()}|{location.lower()}"
            if key in existing_company_title_loc:
                skipped += 1
                continue

        # Map job data to columns
        row_data = {}
        for key, label, width in JOBS_COLUMNS:
            if key == "date_found":
                row_data[key] = date_str
            elif key == "job_id":
                row_data[key] = job_id
            elif key == "match_score":
                row_data[key] = ""  # to be filled by scoring
            elif key == "ats_score":
                row_data[key] = ""
            elif key == "missing_skills":
                row_data[key] = ""
            elif key == "fit_notes":
                row_data[key] = ""
            elif key == "priority":
                row_data[key] = ""
            elif key == "salary_estimate":
                # Autofill from Adzuna real salaries; job JSON may override
                row_data[key] = salary_by_url.get(url, "") or job.get("salary_estimate", "")
            elif key == "status":
                status = job.get("status") or "Not Applied"
                row_data[key] = status
            else:
                # Normalize Source values to lowercase ('LinkedIn'/'Dice' drift)
                val = job.get(key, "")
                if key == "search_mode" and val:
                    val = str(val).strip().lower()
                row_data[key] = val

        # Write row
        for idx, (key, label, width) in enumerate(JOBS_COLUMNS, 1):
            ws.cell(row=next_row, column=idx, value=row_data[key])

        # Format priority cell if set
        if row_data.get("priority") in PRIORITY_COLORS:
            ws.cell(row=next_row, column=_find_col_by_key(JOBS_COLUMNS, "priority")).fill = PRIORITY_COLORS[row_data["priority"]]

        # Apply status color coding to company cell
        status_val = row_data.get("status", "Not Applied")
        if status_val in STATUS_COLORS:
            ws.cell(row=next_row, column=_find_col_by_key(JOBS_COLUMNS, "company")).fill = STATUS_COLORS[status_val]
            if status_val == "Applied":
                ws.cell(row=next_row, column=_find_col_by_key(JOBS_COLUMNS, "company")).font = Font(bold=True)

        next_row += 1
        added += 1
        existing_urls.add(url)
        if job_id:
            existing_job_ids.add(job_id)
        if app_url:
            existing_app_urls.add(app_url)
        if company and title:
            key = f"{company.lower()}|{title.lower()}|{location.lower()}"
            existing_company_title_loc.add(key)

    # Ensure auto-filters on all sheets
    for sheet_name in ['Jobs', 'Applications', 'Resume Versions']:
        ws_filter = wb[sheet_name]
        max_col = ws_filter.max_column
        ws_filter.auto_filter.ref = f'A1:{get_column_letter(max_col)}1'
    
    # Validate all rows before saving
    errors = validate_journal_rows(ws)
    if errors:
        print(f"VALIDATION WARNINGS ({len(errors)}):")
        for e in errors:
            print(f"  {e}")
    
    wb.save(str(path))
    print(f"Added {added} jobs, skipped {skipped} duplicates")
    return added, skipped


def remove_jobs(company=None, title=None, url=None, path=JOURNAL_PATH, all_medlow=False):
    """Remove jobs from the journal by company/title/URL match, or all Medium/Low Not-Applied.

    Args:
        company: Company name to match (case-insensitive substring). Optional.
        title: Optional title filter (case-insensitive substring). Optional.
        url: Optional URL substring to match (e.g. a LinkedIn job ID). Optional.
        path: Path to journal file.
        all_medlow: If True, remove all rows with Priority in (Medium, Low) AND status 'Not Applied'.
    Returns:
        list of removed row labels (str).
    """
    wb = load_workbook(str(path))
    ws = wb['Jobs']
    hdr = [c.value for c in ws[1]]
    def _col(n): return hdr.index(n) + 1 if n in hdr else None
    comp_c, title_c = _col('Company'), _col('Job Title')
    url_c = _col('Job URL')
    prio_c, status_c = _col('Priority'), _col('Status')

    to_delete = []
    for row in ws.iter_rows(min_row=2, max_col=max(len(hdr), 20)):
        r = row[0].row
        if all_medlow:
            p = str(row[prio_c - 1].value or '') if prio_c else ''
            s = str(row[status_c - 1].value or '') if status_c else ''
            if p in ('Medium', 'Low') and s == 'Not Applied':
                to_delete.append(r)
            continue
        row_comp = str(row[comp_c - 1].value or '') if comp_c else ''
        row_title = str(row[title_c - 1].value or '') if title_c else ''
        row_url = str(row[url_c - 1].value or '') if url_c else ''
        match = False
        if company and company.lower() in row_comp.lower():
            if title and title.lower() not in row_title.lower():
                continue
            match = True
        elif url and url in row_url:
            match = True
        elif not company and not url and not all_medlow:
            match = False
        if match:
            to_delete.append(r)

    removed = []
    for r in sorted(to_delete, reverse=True):
        removed.append(f'{ws.cell(r, comp_c).value} — {ws.cell(r, title_c).value}' if comp_c and title_c else f'row {r}')
        ws.delete_rows(r)
    wb.save(str(path))
    return removed


def update_status(company, status, path=JOURNAL_PATH, title=None):
    """Update job status and set Last Updated timestamp. Archives PDFs for Rejected/Closed.
    
    Args:
        company: Company name to match (case-insensitive)
        status: New status (Applied, Interview, Rejected, Closed, Withdrawn, Not Applied)
        title: Optional title filter if multiple roles for same company
        path: Path to journal file
    """
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    import shutil, os
    
    wb = load_workbook(str(path))
    ws = wb['Jobs']
    today = datetime.now().strftime('%Y-%m-%d')
    
    STATUS_COLORS = {
        "Applied": PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        "Interview": PatternFill(start_color='B3D9FF', end_color='B3D9FF', fill_type='solid'),
        "Rejected": PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        "Closed": PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        "Withdrawn": PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }
    
    updated = []
    for row in ws.iter_rows(min_row=2, max_col=20):
        row_company = str(row[JOBS_COL_INDEX["company"] - 1].value) if row[JOBS_COL_INDEX["company"] - 1].value else ''
        row_title = str(row[JOBS_COL_INDEX["title"] - 1].value) if row[JOBS_COL_INDEX["title"] - 1].value else ''
        if company.lower() in row_company.lower():
            if title and title.lower() not in row_title.lower():
                continue
            ws.cell(row=row[0].row, column=JOBS_COL_INDEX["status"]).value = status
            ws.cell(row=row[0].row, column=JOBS_COL_INDEX["last_updated"]).value = today
            if status == 'Applied':
                ws.cell(row=row[0].row, column=JOBS_COL_INDEX["date_applied"]).value = today
            if status in STATUS_COLORS:
                row[JOBS_COL_INDEX["company"] - 1].fill = STATUS_COLORS[status]
                if status == 'Applied':
                    row[JOBS_COL_INDEX["company"] - 1].font = Font(bold=True)
            updated.append(f'{row_company} — {row_title}')
    
    wb.save(str(path))
    
    # Archive PDFs (and review notes) based on status.
    # Applied -> move into applied/; Rejected/Closed/Withdrawn -> move into archived/.
    import re
    def _match(fname):
        c = re.sub(r'[^a-z0-9]', '', company.lower())
        f = re.sub(r'[^a-z0-9]', '', fname.lower())
        return bool(c) and c in f

    resume_dir = os.path.join(os.path.dirname(str(path)), 'resume', 'tailored')
    if status == 'Applied':
        applied_dir = os.path.join(resume_dir, 'applied')
        os.makedirs(applied_dir, exist_ok=True)
        for f in os.listdir(resume_dir):
            if _match(f) and (f.endswith('.pdf') or f.endswith('.txt')):
                shutil.move(os.path.join(resume_dir, f), os.path.join(applied_dir, f))
    elif status in ('Rejected', 'Closed', 'Withdrawn'):
        archive_dir = os.path.join(resume_dir, 'archived')
        os.makedirs(archive_dir, exist_ok=True)
        for f in os.listdir(resume_dir):
            if _match(f) and (f.endswith('.pdf') or f.endswith('.txt')):
                shutil.move(os.path.join(resume_dir, f), os.path.join(archive_dir, f))
        applied_dir = os.path.join(resume_dir, 'applied')
        if os.path.isdir(applied_dir):
            for f in os.listdir(applied_dir):
                if _match(f) and (f.endswith('.pdf') or f.endswith('.txt')):
                    shutil.move(os.path.join(applied_dir, f), os.path.join(archive_dir, f))
    
    for u in updated:
        print(f'{status}: {u}')
    return updated


def validate_journal_rows(ws):
    """Validate journal rows for data integrity. Returns list of error strings."""
    errors = []
    valid_statuses = {"Not Applied", "Applied", "Interview", "Rejected", "Closed", "Withdrawn", None, ""}
    valid_priorities = {"High", "Medium", "Low", None, ""}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        company = str(row[1].value) if row[1].value else ""
        title = str(row[2].value)[:40] if row[2].value else ""
        row_label = f"Row {row_idx} ({company}/{title})"

        # Job ID — should be numeric ID or UUID, never a URL or priority word
        job_id = str(row[JOBS_COL_INDEX["job_id"] - 1].value) if row[JOBS_COL_INDEX["job_id"] - 1].value else ""
        if job_id:
            if job_id.startswith("http"):
                errors.append(f"{row_label}: Job ID contains URL instead of ID — {job_id[:60]}")
            elif job_id in ("Low", "High", "Medium", "New", "Not Applied", "Applied", "Interview", "Rejected"):
                errors.append(f"{row_label}: Job ID contains '{job_id}' — column shift detected")

        # Status — must be a valid status
        status = str(row[JOBS_COL_INDEX["status"] - 1].value) if row[JOBS_COL_INDEX["status"] - 1].value else ""
        if status and status not in valid_statuses:
            errors.append(f"{row_label}: Status='{status}' — not a valid status")

        # Priority — must be High/Medium/Low or empty
        priority = str(row[JOBS_COL_INDEX["priority"] - 1].value) if row[JOBS_COL_INDEX["priority"] - 1].value else ""
        if priority and priority not in valid_priorities:
            errors.append(f"{row_label}: Priority='{priority[:40]}' — not a valid priority level")

        # Date Applied — should be a date or None, not a URL
        date_applied = str(row[JOBS_COL_INDEX["date_applied"] - 1].value) if row[JOBS_COL_INDEX["date_applied"] - 1].value else ""
        if date_applied and date_applied.startswith("http"):
            errors.append(f"{row_label}: Date Applied contains URL instead of date — {date_applied[:60]}")

        # App URL — should be a URL or None, not a status word
        app_url = str(row[JOBS_COL_INDEX["app_url"] - 1].value) if row[JOBS_COL_INDEX["app_url"] - 1].value else ""
        if app_url and app_url in ("New", "Not Applied", "Applied", "Low", "High", "Medium"):
            errors.append(f"{row_label}: App URL='{app_url}' — not a valid URL")

    return errors


def show_status(path=JOURNAL_PATH):
    """Print summary stats from the journal."""
    if not path.exists():
        print("No journal found. Run: python3 journal.py --init")
        return

    wb = load_workbook(str(path), read_only=True)

    print(f"Journal: {path}\n")

    # Jobs sheet
    ws = wb["Jobs"]
    total = ws.max_row - 1
    print(f"Jobs sheet: {total} jobs")

    # Count by source
    sources = {}
    scored = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        mode = row[JOBS_COL_INDEX["search_mode"] - 1] or "unknown"
        sources[mode] = sources.get(mode, 0) + 1
        if row[JOBS_COL_INDEX["match_score"] - 1]:  # match_score
            scored += 1

    for src, count in sorted(sources.items()):
        print(f"  {src}: {count}")
    print(f"  Scored: {scored}/{total}")

    # Applications sheet
    ws = wb["Applications"]
    app_count = ws.max_row - 1
    print(f"\nApplications sheet: {app_count} entries")

    statuses = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        status = row[JOBS_COL_INDEX["status"] - 1] or "No Status"
        statuses[status] = statuses.get(status, 0) + 1
    for st, count in sorted(statuses.items()):
        print(f"  {st}: {count}")

    # Resume sheet
    ws = wb["Resume Versions"]
    resume_count = ws.max_row - 1
    print(f"\nResume Versions sheet: {resume_count} entries")

    wb.close()


def _find_col_by_key(columns, key):
    """Find 1-based column index by key."""
    for idx, (k, label, width) in enumerate(columns, 1):
        if k == key:
            return idx
    return None


def main():
    parser = argparse.ArgumentParser(description="Job Journal — Excel tracking")
    parser.add_argument("--init", action="store_true", help="Create a new journal")
    parser.add_argument("--add", help="Add jobs from JSON file")
    parser.add_argument("--remove", action="store_true", help="Remove jobs matching --company/--title/--url (or all Medium/Low Not-Applied with --remove-all-medlow)")
    parser.add_argument("--company", help="Company name to match when removing")
    parser.add_argument("--title", help="Title filter when removing")
    parser.add_argument("--url", help="Job URL substring to match when removing")
    parser.add_argument("--remove-all-medlow", action="store_true", help="Remove all Medium/Low priority, Not Applied jobs")
    parser.add_argument("--status", action="store_true", help="Show summary stats")
    parser.add_argument("--path", default=str(JOURNAL_PATH), help="Journal file path")
    args = parser.parse_args()

    path = Path(args.path)

    if args.init:
        init_journal(path)
    elif args.add:
        with open(args.add) as f:
            jobs = json.load(f)
        add_jobs(jobs, path)
    elif args.remove or args.remove_all_medlow:
        removed = remove_jobs(company=args.company, title=args.title, url=args.url, path=path,
                              all_medlow=args.remove_all_medlow)
        if removed:
            print(f"Removed {len(removed)} job(s):")
            for r in removed:
                print(f"  - {r}")
        else:
            print("No matching jobs found to remove.")
    elif args.status:
        show_status(path)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()