#!/usr/bin/env python3
"""
Evidence-based Candidate Fit scoring (Stage 2).

For each JD, instead of "does the candidate have keyword X?", this answers
"what evidence does the candidate have for requirement X?" — then weights
required vs preferred and produces a real 0-100 Candidate Fit score.

Pipeline (LLM does semantics, Python does math):
  1. Load the JD's extracted requirements (output/requirements_cache.json).
  2. One LLM call per JD: given the requirements + the candidate's master
     profile (prose), classify each requirement's evidence strength.
  3. Python aggregates: strength -> 1.0/0.8/0.6/0.3/0, weighted by
     required(1.0)/preferred(0.5)/bonus(0.2), to a weighted-average 0-100.

Evidence strengths:
  direct               1.0  — exact capability, recent (~last 2 years)
  older_direct         0.8  — exact capability, 3+ years ago
  strongly_transferable 0.6  — adjacent capability that clearly transfers
  weakly_transferable   0.3  — loosely related
  none                 0.0  — no evidence

Result cached in output/evidence_cache.json, keyed by job URL:
  {
    "<url>": {
      "engine": "ollama",
      "extracted_ts": "...",
      "candidate_fit": 78,
      "evidence": [
        {"index": 1, "text": "...", "type": "required",
         "evidence_strength": "direct", "evidence_quote": "...", "confidence": 0.9}
      ]
    }
  }

Usage:
  python3 scripts/evidence_score.py                 # score all pending JDs (no cap)
  python3 scripts/evidence_score.py --backfill 200  # process up to 200 uncached
  python3 scripts/evidence_score.py --url <job_url> # one JD (re-scores)
  python3 scripts/evidence_score.py --report        # cache stats

Config via env (same as extract_requirements.py):
  EMR_ENDPOINT, EMR_MODEL, EMR_TIMEOUT, EMR_MAX_TOKENS, EMR_WORKERS,
  EMR_MAX_JD_CHARS, OLLAMA_API_KEY (falls back to ~/.hermes/.env)

Standard library only.
"""

import argparse
import json
import os
import re
import subprocess
import sys
import threading
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent.parent
REQS_PATH = BASE_DIR / "output" / "requirements_cache.json"
EVID_PATH = BASE_DIR / "output" / "evidence_cache.json"
PROFILE_PATH = BASE_DIR / "resume" / "MASTER_PROFILE.md"
JOURNAL_PATH = BASE_DIR / "journal.xlsx"

ENDPOINT = os.environ.get("EMR_ENDPOINT", "https://ollama.com/v1")
MODEL = os.environ.get("EMR_MODEL", "glm-5.3-flash")
TIMEOUT_S = int(os.environ.get("EMR_TIMEOUT", "120"))
MAX_TOKENS = int(os.environ.get("EMR_MAX_TOKENS", "12000"))
MAX_WORKERS = int(os.environ.get("EMR_WORKERS", "5"))
MAX_JD_CHARS = int(os.environ.get("EMR_MAX_JD_CHARS", "12000"))

# Built by concatenation so source files never contain literal think-tags
CLOSE_THINK = "<" + "/think>"

_CACHE_LOCK = threading.Lock()

# Evidence strength -> score
STRENGTH_SCORE = {
    "direct": 1.0,
    "older_direct": 0.8,
    "strongly_transferable": 0.6,
    "weakly_transferable": 0.3,
    "none": 0.0,
}
# Requirement type -> weight
TYPE_WEIGHT = {"required": 1.0, "preferred": 0.5, "bonus": 0.2}


def _load_api_key():
    v = os.environ.get("OLLAMA_API_KEY")
    if v:
        return v
    try:
        for line in (Path.home() / ".hermes" / ".env").read_text().splitlines():
            if line.strip().startswith("OLLAMA_API_KEY="):
                return line.partition("=")[2].strip().strip('"').strip("'")
    except Exception:
        pass
    return ""


API_KEY = _load_api_key()


def _ollama_generate(prompt):
    """Call an OpenAI-compatible chat endpoint. Returns (parsed_json, error)."""
    payload = json.dumps({
        "model": MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "max_tokens": MAX_TOKENS,
    })
    headers = ["-H", "Content-Type: application/json"]
    if API_KEY:
        headers += ["-H", "Authorization: Bearer " + API_KEY]
    try:
        r = subprocess.run(
            ["curl", "-s", "--max-time", str(TIMEOUT_S), *headers,
             "-d", payload, ENDPOINT.rstrip("/") + "/chat/completions"],
            capture_output=True, text=True, timeout=TIMEOUT_S + 10)
        if not r.stdout.strip():
            return None, f"empty response (curl exit {r.returncode})"
        try:
            data = json.loads(r.stdout)
        except json.JSONDecodeError:
            return None, "non-JSON response: " + r.stdout[:150]
        if data.get("error"):
            return None, "API error: " + str(data["error"])[:200]
        msg = (data.get("choices") or [{}])[0].get("message", {})
        raw = (msg.get("content") or msg.get("reasoning_content")
               or msg.get("thinking") or msg.get("reasoning") or "").strip()
        if not raw:
            return None, "no content in response: " + r.stdout[:300]
        if CLOSE_THINK in raw:
            raw = raw.split(CLOSE_THINK)[-1].strip()
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip()).strip()
        try:
            return json.loads(raw), None
        except json.JSONDecodeError:
            salvaged = _salvage_evidence_json(raw)
            if salvaged is not None:
                return salvaged, None
            return None, "unparseable model output: " + raw[:150]
    except subprocess.TimeoutExpired:
        return None, f"timeout after {TIMEOUT_S}s"
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def _salvage_evidence_json(raw):
    """Salvage evidence objects from truncated/invalid JSON.

    Evidence objects are flat (no nested braces), so individually parsing
    every {...} span recovers complete items even when the model output was
    cut off mid-array by a token limit. Returns {"evidence": [...]} or None.
    """
    objs = []
    for m in re.finditer(r"\{[^{}]*\}", raw, re.S):
        try:
            obj = json.loads(m.group(0))
        except json.JSONDecodeError:
            continue
        if isinstance(obj, dict) and obj.get("index") is not None:
            objs.append(obj)
    return {"evidence": objs} if objs else None


def _build_prompt(requirements, profile):
    """Build the evidence-scoring prompt for one JD."""
    lines = []
    for i, r in enumerate(requirements, 1):
        lines.append(f"{i}. [{r['type']}] ({r['category']}) {r['text']}")
    req_block = "\n".join(lines)
    return f"""You are scoring a candidate's fit against a job's requirements using evidence from their resume.

CANDIDATE PROFILE:
<<<
{profile[:MAX_JD_CHARS]}
>>>

JOB REQUIREMENTS (numbered):
{req_block}

For EACH requirement, return a JSON object with:
- "index": the requirement number (int)
- "evidence_strength": one of "direct" | "older_direct" | "strongly_transferable" | "weakly_transferable" | "none"
- "evidence_quote": a short verbatim quote from the profile supporting it (or "" if none)
- "confidence": 0.0-1.0

Rules:
- "direct": the profile shows this exact capability recently (last ~2 years).
- "older_direct": direct but from 3+ years ago.
- "strongly_transferable": adjacent capability that clearly transfers.
- "weakly_transferable": loosely related.
- "none": no evidence. Be honest — if there's no evidence, say "none".

Output ONLY a JSON object: {{"evidence": [{{"index": 1, "evidence_strength": "...", "evidence_quote": "...", "confidence": 0.9}}, ...]}}"""


def _compute_fit(requirements, evidence_list):
    """Aggregate evidence into a 0-100 weighted-average Candidate Fit score."""
    by_index = {e.get("index"): e for e in evidence_list if isinstance(e, dict)}
    total_weight = 0.0
    weighted = 0.0
    for i, req in enumerate(requirements, 1):
        w = TYPE_WEIGHT.get(req.get("type", "required"), 1.0)
        total_weight += w
        ev = by_index.get(i, {})
        strength = str(ev.get("evidence_strength", "none")).lower()
        s = STRENGTH_SCORE.get(strength, 0.0)
        weighted += s * w
    if total_weight == 0:
        return 0
    return round(weighted / total_weight * 100)


def score_one(url, requirements, profile, cache):
    """Score one JD, update cache. Returns (candidate_fit, engine, error).
    Retries once if the first LLM call yields no parseable evidence (the
    thinking model is nondeterministic and sometimes burns its budget on
    reasoning without emitting JSON)."""
    rec = {"url": url, "extracted_ts": datetime.now().isoformat(),
           "candidate_fit": None, "evidence": []}
    data, err = _ollama_generate(_build_prompt(requirements, profile))
    for _ in range(2):
        if data and isinstance(data.get("evidence"), list):
            break
        # Retry — the thinking model is nondeterministic and sometimes burns
        # its budget on reasoning without emitting JSON
        data, err = _ollama_generate(_build_prompt(requirements, profile))
    if data and isinstance(data.get("evidence"), list):
        rec["evidence"] = data["evidence"]
        rec["candidate_fit"] = _compute_fit(requirements, data["evidence"])
        rec["engine"] = "ollama"
    else:
        rec["engine"] = "error"
        rec["llm_error"] = err or "no evidence returned"
    with _CACHE_LOCK:
        cache[url] = rec
        save_cache(cache)
    return rec["candidate_fit"], rec["engine"], rec.get("llm_error")


def load_cache():
    try:
        with open(EVID_PATH) as f:
            return json.load(f)
    except Exception:
        return {}


def save_cache(cache):
    with open(EVID_PATH, "w") as f:
        json.dump(cache, f, indent=2)


def _write_journal_fit(url, fit):
    """Write candidate_fit to the journal column for the matching row."""
    from openpyxl import load_workbook
    from journal import JOBS_COL_INDEX, _style_header_cell
    wb = load_workbook(str(JOURNAL_PATH))
    ws = wb["Jobs"]
    url_col = JOBS_COL_INDEX["url"]
    fit_col = JOBS_COL_INDEX["candidate_fit"]
    # Ensure header exists (append-only column), styled like the rest
    if ws.cell(row=1, column=fit_col).value in (None, ""):
        _style_header_cell(ws.cell(row=1, column=fit_col, value="Candidate Fit"))
    for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col):
        if str(row[0].value or "").strip() == url:
            ws.cell(row=row[0].row, column=fit_col, value=fit)
            break
    wb.save(str(JOURNAL_PATH))
    wb.close()


def _write_journal_fits(cache, urls):
    """Write candidate_fit for a list of URLs to the journal in one pass."""
    from openpyxl import load_workbook
    from journal import JOBS_COL_INDEX, _style_header_cell
    wb = load_workbook(str(JOURNAL_PATH))
    ws = wb["Jobs"]
    url_col = JOBS_COL_INDEX["url"]
    fit_col = JOBS_COL_INDEX["candidate_fit"]
    if ws.cell(row=1, column=fit_col).value in (None, ""):
        _style_header_cell(ws.cell(row=1, column=fit_col, value="Candidate Fit"))
    url_to_fit = {u: cache[u].get("candidate_fit") for u in urls
                  if cache.get(u, {}).get("candidate_fit") is not None}
    if not url_to_fit:
        wb.close()
        return
    for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col):
        u = str(row[0].value or "").strip()
        if u in url_to_fit:
            ws.cell(row=row[0].row, column=fit_col, value=url_to_fit[u])
    wb.save(str(JOURNAL_PATH))
    wb.close()
    print(f"Wrote candidate_fit to journal for {len(url_to_fit)} job(s)")


def main():
    ap = argparse.ArgumentParser(description="Evidence-based Candidate Fit scoring")
    ap.add_argument("--backfill", type=int, default=0, help="max JDs to score (0 = all pending, no cap)")
    ap.add_argument("--url", help="score a single JD by URL (re-scores)")
    ap.add_argument("--report", action="store_true", help="print cache stats")
    args = ap.parse_args()

    try:
        with open(REQS_PATH) as f:
            reqs_cache = json.load(f)
    except Exception as e:
        print(f"ERROR: cannot read {REQS_PATH}: {e}")
        sys.exit(1)
    try:
        profile = open(PROFILE_PATH).read()
    except Exception as e:
        print(f"ERROR: cannot read {PROFILE_PATH}: {e}")
        sys.exit(1)

    cache = load_cache()

    if args.report:
        total = len(cache)
        fits = [v.get("candidate_fit") for v in cache.values() if v.get("candidate_fit") is not None]
        engines = Counter(v.get("engine", "?") for v in cache.values())
        print(f"Scored JDs: {total}/{len(reqs_cache)}")
        print(f"Engines: {dict(engines)}")
        if fits:
            print(f"Avg candidate_fit: {sum(fits)/len(fits):.1f} | min {min(fits)} | max {max(fits)}")
        return

    if args.url:
        reqs = (reqs_cache.get(args.url) or {}).get("requirements", [])
        if not reqs:
            print(f"ERROR: no extracted requirements for {args.url}")
            sys.exit(1)
        fit, engine, err = score_one(args.url, reqs, profile, cache)
        suffix = f" | error: {err}" if err else ""
        print(f"{engine}: candidate_fit={fit} — {args.url}{suffix}")
        _write_journal_fit(args.url, fit)
        return

    # Batch mode
    cap = args.backfill if args.backfill > 0 else None  # None = no cap
    todo = [u for u, rec in reqs_cache.items()
            if rec.get("requirements") and u not in cache][:cap]
    print(f"Scoring {len(todo)} JD(s) with {MODEL} ({MAX_WORKERS} workers)...")
    engines = Counter()
    errors = Counter()
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(score_one, u, reqs_cache[u]["requirements"], profile, cache): u
                   for u in todo}
        for fut in as_completed(futures):
            fit, engine, err = fut.result()
            engines[engine] += 1
            if err:
                errors[err[:100]] += 1
            done += 1
            if done % 5 == 0 or done == len(todo):
                print(f"  {done}/{len(todo)} done (engines: {dict(engines)})", flush=True)
    print(f"Done. engines: {dict(engines)} | cache size: {len(cache)}")
    if errors:
        print("LLM errors (first 3):")
        for e, n in list(errors.items())[:3]:
            print(f"  {n}x {e}")

    # Write candidate_fit to the journal for all scored JDs (single serialized pass)
    _write_journal_fits(cache, todo)


if __name__ == "__main__":
    main()